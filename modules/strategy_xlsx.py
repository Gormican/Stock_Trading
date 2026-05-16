"""
strategy_xlsx.py — Parse Strategies.xlsx into rich strategy dicts.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore


class OpenpyxlMissing(RuntimeError):
    """Raised when openpyxl isn't installed but xlsx parsing was attempted."""


def _require_openpyxl() -> None:
    if openpyxl is None:
        raise OpenpyxlMissing(
            "openpyxl is not installed. Run: pip install openpyxl"
        )


# Column indexes (1-based) in each strategy sheet
COL_ITEM           = 1   # A
COL_CATEGORY_IN    = 2   # B  (yellow input)
COL_SUBCAT_IN      = 4   # D  (yellow input)
COL_COMPONENT_IN   = 7   # G  (yellow input)
COL_HIGHER_BETTER  = 10  # J
COL_X              = 11  # K
COL_Y              = 12  # L
COL_Z              = 13  # M
COL_SOURCE         = 18  # R
COL_DATA_KEY       = 19  # S
COL_NOTES          = 20  # T


def _read_strategy_index(wb) -> list[dict]:
    """Return the rows of the Strategies index sheet."""
    if "Strategies" not in wb.sheetnames:
        return []
    ws = wb["Strategies"]
    out: list[dict] = []
    header_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value and str(ws.cell(r, 1).value).strip().lower() == "active":
            header_row = r
            break
    if header_row is None:
        return out
    for r in range(header_row + 1, ws.max_row + 1):
        active = ws.cell(r, 1).value
        name   = ws.cell(r, 2).value
        sheet  = ws.cell(r, 3).value
        desc   = ws.cell(r, 4).value
        mod    = ws.cell(r, 5).value
        if not name or not sheet:
            continue
        out.append({
            "active":        str(active).strip().upper() == "YES",
            "name":          str(name).strip(),
            "sheet":         str(sheet).strip(),
            "description":   "" if desc is None else str(desc),
            "last_modified": "" if mod is None else str(mod),
        })
    return out


def _f(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _yes(v) -> bool:
    if v is None:
        return False
    return str(v).strip().lower() in ("y", "yes", "true", "1")


def _read_strategy_sheet(ws) -> dict:
    """Walk a strategy sheet and build the 3-level tree."""
    categories: list[dict] = []
    cur_cat: dict | None = None
    cur_sub: dict | None = None

    header_row = None
    for r in range(1, min(ws.max_row, 10) + 1):
        if ws.cell(r, COL_ITEM).value and str(ws.cell(r, COL_ITEM).value).strip().lower() == "item":
            header_row = r
            break
    if header_row is None:
        return {"categories": []}

    for r in range(header_row + 1, ws.max_row + 1):
        item = ws.cell(r, COL_ITEM).value
        if item is None or str(item).strip() == "":
            continue
        item_s = str(item).strip()
        if item_s.lower() == "totals":
            continue

        cat_in  = _f(ws.cell(r, COL_CATEGORY_IN).value)
        sub_in  = _f(ws.cell(r, COL_SUBCAT_IN).value)
        comp_in = _f(ws.cell(r, COL_COMPONENT_IN).value)
        data_key = ws.cell(r, COL_DATA_KEY).value

        if cat_in is not None:
            cur_cat = {"name": item_s, "weight_input": cat_in, "subcategories": []}
            categories.append(cur_cat)
            cur_sub = None
            continue

        if sub_in is not None:
            cur_sub = {"name": item_s, "weight_input": sub_in, "components": []}
            if cur_cat is None:
                continue
            cur_cat["subcategories"].append(cur_sub)
            if data_key:
                # Merged subcategory = single-component subcat
                cur_sub["components"].append({
                    "name": item_s,
                    "weight_input": 1.0,
                    "higher_better": _yes(ws.cell(r, COL_HIGHER_BETTER).value),
                    "x": _f(ws.cell(r, COL_X).value),
                    "y": _f(ws.cell(r, COL_Y).value),
                    "z": _f(ws.cell(r, COL_Z).value),
                    "source":   str(ws.cell(r, COL_SOURCE).value or "").strip(),
                    "data_key": str(data_key).strip(),
                    "notes":    str(ws.cell(r, COL_NOTES).value or ""),
                })
            continue

        if comp_in is not None:
            if cur_sub is None:
                continue
            cur_sub["components"].append({
                "name": item_s,
                "weight_input": comp_in,
                "higher_better": _yes(ws.cell(r, COL_HIGHER_BETTER).value),
                "x": _f(ws.cell(r, COL_X).value),
                "y": _f(ws.cell(r, COL_Y).value),
                "z": _f(ws.cell(r, COL_Z).value),
                "source":   str(ws.cell(r, COL_SOURCE).value or "").strip(),
                "data_key": str(data_key).strip() if data_key else "",
                "notes":    str(ws.cell(r, COL_NOTES).value or ""),
            })

    return {"categories": categories}


def normalize_tree(tree: dict) -> dict:
    """Add normalized `weight` and `pct_of_gpa` fields at every level."""
    cats = tree.get("categories", [])
    cat_total = sum(c.get("weight_input") or 0 for c in cats) or 1.0
    for c in cats:
        c["weight"] = (c.get("weight_input") or 0) / cat_total
        subs = c.get("subcategories", [])
        sub_total = sum(s.get("weight_input") or 0 for s in subs) or 1.0
        for s in subs:
            s["weight"] = (s.get("weight_input") or 0) / sub_total
            comps = s.get("components", [])
            comp_total = sum(p.get("weight_input") or 0 for p in comps) or 1.0
            for p in comps:
                p["weight"] = (p.get("weight_input") or 0) / comp_total
                p["pct_of_gpa"] = c["weight"] * s["weight"] * p["weight"]
    return tree


_CAT_LEGACY = {
    "sentiment":    "sentiment",
    "fundamentals": "fundamentals",
    "technical":    "technical",
}

_SUB_LEGACY = {
    "valuation":          ("fund_sub", "valuation"),
    "financials":         ("fund_sub", "financial"),
    "financial":          ("fund_sub", "financial"),
    "estimates":          ("fund_sub", "estimates"),
    "analyst buy rating": ("fund_sub", "estimates"),
    "trend":              ("tech_sub", "trend"),
    "oscillators":        ("tech_sub", "oscillators"),
}


def to_legacy_weights(tree: dict) -> dict:
    """Collapse the rich tree into the dict shape GPAEngine expects."""
    tree = normalize_tree({"categories": [
        {**c, "subcategories": [{**s, "components": list(s.get("components", []))}
                                for s in c.get("subcategories", [])]}
        for c in tree.get("categories", [])
    ]})

    out = {
        "sentiment":    0.0,
        "fundamentals": 0.0,
        "technical":    0.0,
        "fund_sub":     {"valuation": 0.0, "financial": 0.0, "estimates": 0.0},
        "tech_sub":     {"trend":     0.0, "oscillators":   0.0},
    }
    for c in tree["categories"]:
        cat_key = _CAT_LEGACY.get(c["name"].strip().lower())
        if cat_key is None:
            continue
        out[cat_key] = c["weight"]
        if cat_key == "fundamentals":
            for s in c.get("subcategories", []):
                k = _SUB_LEGACY.get(s["name"].strip().lower())
                if k and k[0] == "fund_sub":
                    out["fund_sub"][k[1]] += s["weight"]
        elif cat_key == "technical":
            for s in c.get("subcategories", []):
                k = _SUB_LEGACY.get(s["name"].strip().lower())
                if k and k[0] == "tech_sub":
                    out["tech_sub"][k[1]] += s["weight"]

    def _norm(d):
        t = sum(d.values())
        return {k: (v / t if t else 0) for k, v in d.items()}
    out["fund_sub"] = _norm(out["fund_sub"])
    out["tech_sub"] = _norm(out["tech_sub"])
    return out


def load_xlsx(path: str | Path) -> dict[str, dict]:
    """Parse Strategies.xlsx. Raises OpenpyxlMissing if openpyxl isn't installed."""
    _require_openpyxl()
    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    index = _read_strategy_index(wb)

    result: dict[str, Any] = {"_index": index, "_active": None}
    for entry in index:
        sheet_name = entry["sheet"]
        if sheet_name not in wb.sheetnames:
            continue
        tree = _read_strategy_sheet(wb[sheet_name])
        normalize_tree(tree)
        legacy = to_legacy_weights(tree)
        result[entry["name"]] = {
            "name":           entry["name"],
            "active":         entry["active"],
            "description":    entry["description"],
            "last_modified":  entry["last_modified"],
            "sheet":          sheet_name,
            "tree":           tree,
            "legacy_weights": legacy,
        }
        if entry["active"]:
            result["_active"] = entry["name"]
    return result


def get_known_data_keys() -> set[str]:
    """Data keys the existing pipeline already fetches."""
    return {
        "av_score", "av_news_sentiment_score",
        "trailingPE", "pe_ttm",
        "pegRatio", "peg",
        "enterpriseToEbitda", "ev_ebitda",
        "returnOnEquity", "roe",
        "debtToEquity", "debt_equity",
        "dividendYield", "dividend_yield",
        "earningsGrowth", "earnings_growth_yoy",
        "recommendationMean", "analyst_recommendation",
        "earnings_history_beat_pct", "beat_rate", "beat_expectations",
        "trend_3mo_slope_pct", "trend_1yr_slope_pct",
        "price_vs_sma50_pct", "price_vs_sma200_pct",
        "macd_hist_value", "stoch_k", "rsi_14",
    }


def grade_value(value, comp: dict):
    """Apply x/y/z thresholds; return (grade, score) or (None, None)."""
    if value is None or comp.get("x") is None or comp.get("y") is None or comp.get("z") is None:
        return None, None
    x, y, z = comp["x"], comp["y"], comp["z"]
    if comp.get("higher_better"):
        if value >= x: return "A", 4
        if value >= y: return "B", 3
        if value >= z: return "C", 2
        return "D", 1
    else:
        if value <= x: return "A", 4
        if value <= y: return "B", 3
        if value <= z: return "C", 2
        return "D", 1
